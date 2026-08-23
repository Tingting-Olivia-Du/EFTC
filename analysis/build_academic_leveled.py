"""
Applies the six-level Bauer & Nation (1993) affix expansion to the three
academic word lists (AWL, NAWL, AVL), so that they are analysed at the
same counting unit as the general word lists (HSWL, CET-4, CET-6) rather
than in their published form.

WHY THIS IS NEEDED
------------------
The three lists are not published at the same counting unit:

  AWL   570 families, members include derivations   -> approx. Level 6
  NAWL  963 headwords, inflections/variants only    -> approx. Level 2
  AVL   1,991 families, members include derivations -> approx. Level 6

Comparing their published forms therefore confounds lexical selection with
counting unit: the NAWL's low standalone coverage is partly an artefact of
its being a lemma list. Rebuilding all three from their headwords through
the identical pipeline removes that confound.

METHOD
------
Level 1 is obtained by table mapping -- taking the headword key of each
list's own nested structure, which recovers exactly the published family/
headword counts (570 / 963 / 1,991, verified). Levels 2-6 are then built
with the same LevelListHandler used for the general lists: the AntBNC
lemma database for Level 2, the BNC/COCA word-family database for Level 6,
and the Bauer & Nation affix rules (with the corrected Level-3 prefix
handling) for Levels 3-5.

For the NAWL this also supplies the levels its published form cannot
express: distributed as lemmas, it has no derivational members of its own,
so Levels 3-6 are unavailable from the list itself and are recovered here
via the shared family database.

Run (from repo root): python3 analysis/build_academic_leveled.py
"""
import csv
import json
from pathlib import Path

import openpyxl

from coverage_lib import Corpus, coverage_stats
from java_port_pipeline import (LevelListHandler, AffixLevelHandlerFixed,
                                 build_family_lookup)

ROOT = Path(__file__).resolve().parent.parent
WL = ROOT / "data" / "wordlists"
OUT = ROOT / "output"
OUT.mkdir(exist_ok=True)

PUBLISHED_HEADWORDS = {"AWL": 570, "NAWL": 963, "AVL": 1991}


# ---------------------------------------------------------------------------
# Level-1 extraction by table mapping, one reader per source format
# ---------------------------------------------------------------------------
def awl_headwords():
    """AWL_nested.json: {sublist: {headword: {"subwords": [...]}}}"""
    data = json.load(open(ROOT / "archive/wordlist/AWL_nested.json", encoding="utf-8"))
    return {h.lower() for sublist in data.values() for h in sublist}


def nawl_headwords():
    """NAWL_nested.json: {headword: [inflected forms]} -- keys are the lemmas.

    This is the mapping the published list already encodes: its 963 keys are
    Level 1, and key + values together are its Level-2 (lemma) form.
    """
    data = json.load(open(ROOT / "archive/wordlist/NAWL_nested.json", encoding="utf-8"))
    return {h.lower() for h in data}


def avl_headwords():
    """families-AVL.xlsx 'data' sheet: the 'family' column is the family head."""
    wb = openpyxl.load_workbook(WL / "families-AVL.xlsx", read_only=True)
    rows = list(wb["data"].iter_rows(values_only=True))[1:]
    return {str(r[1]).lower() for r in rows}


READERS = {"AWL": awl_headwords, "NAWL": nawl_headwords, "AVL": avl_headwords}


def published_forms(name):
    """The list's own published word-form set, for reference comparison."""
    if name == "AWL":
        data = json.load(open(ROOT / "archive/wordlist/AWL_nested.json", encoding="utf-8"))
        out = set()
        for sublist in data.values():
            for h, info in sublist.items():
                out.add(h.lower())
                out.update(m.lower() for m in (info.get("subwords") or []))
        return out
    if name == "NAWL":
        data = json.load(open(ROOT / "archive/wordlist/NAWL_nested.json", encoding="utf-8"))
        out = set()
        for h, members in data.items():
            out.add(h.lower())
            out.update(m.lower() for m in (members or []))
        return out
    wb = openpyxl.load_workbook(WL / "families-AVL.xlsx", read_only=True)
    rows = list(wb["data"].iter_rows(values_only=True))[1:]
    return {str(r[3]).lower() for r in rows}


def main():
    corpus = Corpus(WL / "eftc_corpus.json")
    family_db, family_reverse = build_family_lookup()
    lemma_db = json.load(open(WL / "antbnc_lemma_database.json", encoding="utf-8"))
    affix = AffixLevelHandlerFixed()
    supplement = {w.lower() for w in json.load(open(WL / "bnc_coca_supplement_combined.json"))}

    # General Composite (Level 6) + supplement -- the baseline the starred rows use
    gc = set()
    for n in ("HSWL", "CET4", "CET6"):
        gc |= set(json.load(open(WL / "leveled_java_port" / n / "level6.json")))
    gcp = gc | supplement

    rows, starred = [], []
    print(f"{'List':5s}{'Lvl':>4s}{'size':>8s}{'attest%':>9s}{'type%':>8s}{'token%':>8s}"
          f"{'   |':>4s}{'+supp tok%':>11s}{'starred tok%':>13s}")
    for name, reader in READERS.items():
        level1 = reader()
        assert len(level1) == PUBLISHED_HEADWORDS[name], \
            f"{name}: extracted {len(level1)} headwords, expected {PUBLISHED_HEADWORDS[name]}"

        handler = LevelListHandler(level1, family_db, family_reverse, lemma_db, affix)
        levels = {1: handler.level1, 2: handler.level2, 3: handler.level3,
                  4: handler.level4, 5: handler.level5, 6: handler.level6}

        out_dir = WL / "leveled_academic" / name
        out_dir.mkdir(parents=True, exist_ok=True)
        for lvl in range(1, 7):
            words = levels[lvl].all_words()
            json.dump(sorted(words), open(out_dir / f"level{lvl}.json", "w", encoding="utf-8"),
                      ensure_ascii=False, indent=1)

            s = coverage_stats(words, corpus)
            s_supp = coverage_stats(words | supplement, corpus)
            s_star = coverage_stats(gcp | words, corpus)
            rows.append({"list": name, "level": lvl, "size": s["wordlist_size"],
                          "attestation_pct": round(s["attestation_rate_pct"], 2),
                          "type_pct": round(s["type_coverage_pct"], 2),
                          "token_pct": round(s["token_coverage_pct"], 2),
                          "plus_supp_size": s_supp["wordlist_size"],
                          "plus_supp_token_pct": round(s_supp["token_coverage_pct"], 2)})
            starred.append({"list": f"{name}*", "level": lvl, "size": s_star["wordlist_size"],
                             "attestation_pct": round(s_star["attestation_rate_pct"], 2),
                             "type_pct": round(s_star["type_coverage_pct"], 2),
                             "token_pct": round(s_star["token_coverage_pct"], 2)})
            print(f"{name:5s}{lvl:4d}{s['wordlist_size']:8d}{s['attestation_rate_pct']:9.2f}"
                  f"{s['type_coverage_pct']:8.2f}{s['token_coverage_pct']:8.2f}{'   |':>4s}"
                  f"{s_supp['token_coverage_pct']:11.2f}{s_star['token_coverage_pct']:13.2f}")

        pub = published_forms(name)
        rebuilt6 = levels[6].all_words()
        print(f"      published form: {len(pub):5d} forms   rebuilt Level 6: {len(rebuilt6):5d}   "
              f"shared: {len(pub & rebuilt6):5d}  "
              f"(published token% = {coverage_stats(pub, corpus)['token_coverage_pct']:.2f})")

    with open(OUT / "academic_leveled.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys())); w.writeheader(); w.writerows(rows)
    with open(OUT / "academic_leveled_starred.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(starred[0].keys())); w.writeheader(); w.writerows(starred)
    print(f"\nWrote {OUT/'academic_leveled.csv'} and {OUT/'academic_leveled_starred.csv'}")
    print(f"Word lists written to {WL/'leveled_academic'}/{{AWL,NAWL,AVL}}/level{{1..6}}.json")


if __name__ == "__main__":
    main()
